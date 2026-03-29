from django.contrib.auth.decorators import login_required
from django.db import models as db_models
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from .models import Contact, Interaction, Task


PAGE_SIZE = 20


@login_required
def contact_list(request):
    contacts = Contact.objects.filter(fc=request.user)

    q = request.GET.get("q", "").strip()
    if q:
        contacts = contacts.filter(
            db_models.Q(name__icontains=q) | db_models.Q(company_name__icontains=q)
        )

    industry = request.GET.get("industry", "").strip()
    if industry:
        contacts = contacts.filter(industry=industry)

    page = int(request.GET.get("page", 1))
    offset = (page - 1) * PAGE_SIZE
    page_contacts = contacts[offset : offset + PAGE_SIZE]
    has_more = contacts[offset + PAGE_SIZE : offset + PAGE_SIZE + 1].exists()

    # Search/filter or infinite scroll → return list items only
    if request.htmx and request.htmx.target == "contact-list":
        return render(request, "contacts/partials/contact_list_items.html", {
            "contacts": page_contacts, "q": q, "industry": industry,
            "page": page, "has_more": has_more,
        })
    # Infinite scroll "load more" → return just the new items
    if request.htmx and request.GET.get("page"):
        return render(request, "contacts/partials/contact_list_items.html", {
            "contacts": page_contacts, "q": q, "industry": industry,
            "page": page, "has_more": has_more, "append": True,
        })

    total_count = contacts.count()
    return render(request, "contacts/contact_list.html", {
        "contacts": page_contacts,
        "q": q,
        "industry": industry,
        "page": 1,
        "has_more": has_more,
        "total_count": total_count,
    })


@login_required
def contact_create(request):
    if request.method == "POST":
        contact = Contact.objects.create(
            fc=request.user,
            name=request.POST["name"],
            phone=request.POST.get("phone", ""),
            company_name=request.POST.get("company_name", ""),
            industry=request.POST.get("industry", ""),
            region=request.POST.get("region", ""),
            revenue_range=request.POST.get("revenue_range", ""),
            employee_count=request.POST.get("employee_count") or None,
            memo=request.POST.get("memo", ""),
        )
        if request.htmx:
            return HttpResponse(
                status=204,
                headers={"HX-Redirect": f"/contacts/{contact.pk}/"},
            )
        return redirect("contacts:detail", pk=contact.pk)

    template = "contacts/partials/contact_form_content.html" if request.htmx else "contacts/contact_form.html"
    return render(request, template, {"editing": False})


INTERACTION_PAGE_SIZE = 15


@login_required
def contact_detail(request, pk):
    contact = get_object_or_404(Contact, pk=pk, fc=request.user)

    page = int(request.GET.get("ipage", 1))
    offset = (page - 1) * INTERACTION_PAGE_SIZE
    interactions = contact.interactions.all()[offset : offset + INTERACTION_PAGE_SIZE]
    has_more_interactions = contact.interactions.all()[offset + INTERACTION_PAGE_SIZE : offset + INTERACTION_PAGE_SIZE + 1].exists()
    latest_brief = contact.briefs.first()

    # Infinite scroll for interactions
    if request.htmx and request.GET.get("ipage"):
        return render(request, "contacts/partials/interaction_timeline.html", {
            "contact": contact,
            "interactions": interactions,
            "ipage": page,
            "has_more_interactions": has_more_interactions,
            "append": True,
        })

    template = "contacts/partials/contact_detail_content.html" if request.htmx else "contacts/contact_detail.html"
    return render(request, template, {
        "contact": contact,
        "interactions": interactions,
        "latest_brief": latest_brief,
        "ipage": 1,
        "has_more_interactions": has_more_interactions,
    })


@login_required
def contact_edit(request, pk):
    contact = get_object_or_404(Contact, pk=pk, fc=request.user)

    if request.method == "POST":
        contact.name = request.POST["name"]
        contact.phone = request.POST.get("phone", "")
        contact.company_name = request.POST.get("company_name", "")
        contact.industry = request.POST.get("industry", "")
        contact.region = request.POST.get("region", "")
        contact.revenue_range = request.POST.get("revenue_range", "")
        contact.employee_count = request.POST.get("employee_count") or None
        contact.memo = request.POST.get("memo", "")
        contact.save()

        if request.htmx:
            return HttpResponse(
                status=204,
                headers={"HX-Redirect": f"/contacts/{contact.pk}/"},
            )
        return redirect("contacts:detail", pk=contact.pk)

    template = "contacts/partials/contact_form_content.html" if request.htmx else "contacts/contact_form.html"
    return render(request, template, {"contact": contact, "editing": True})


@login_required
def contact_delete(request, pk):
    contact = get_object_or_404(Contact, pk=pk, fc=request.user)
    if request.method == "POST":
        contact.delete()
        if request.htmx:
            return HttpResponse(
                status=204,
                headers={"HX-Redirect": "/contacts/"},
            )
        return redirect("contacts:list")
    return redirect("contacts:detail", pk=pk)


@login_required
def contact_delete_all(request):
    """DEV ONLY: Delete all contacts for current user."""
    from django.conf import settings as django_settings

    if not django_settings.DEBUG:
        return HttpResponse(status=403)

    if request.method == "POST":
        count = Contact.objects.filter(fc=request.user).count()
        Contact.objects.filter(fc=request.user).delete()
        if request.htmx:
            return HttpResponse(
                status=204,
                headers={"HX-Redirect": "/contacts/"},
            )
        return redirect("contacts:list")

    count = Contact.objects.filter(fc=request.user).count()
    return render(request, "contacts/delete_all_confirm.html", {"count": count})


@login_required
def contact_search(request):
    q = request.GET.get("q", "").strip()
    contacts = Contact.objects.filter(fc=request.user)
    if q:
        contacts = contacts.filter(
            db_models.Q(name__icontains=q) | db_models.Q(company_name__icontains=q)
        )[:10]
    return render(request, "contacts/partials/search_results.html", {
        "contacts": contacts,
    })


@login_required
def contact_import(request):
    """Step 1: Upload Excel/CSV file."""
    if request.method == "POST" and request.FILES.get("file"):
        return _import_analyze(request)

    return render(request, "contacts/import.html")


MAX_IMPORT_FILE_SIZE = 5 * 1024 * 1024  # 5MB (shapes can bloat files)
MAX_IMPORT_ROWS = 1000


def _preprocess_excel(src_path):
    """Strip drawings/shapes from Excel, keep cell data only.

    Bloated files with hidden shape objects (e.g. 59,000 empty shapes from
    copy-paste) cause severe performance issues. This creates a clean copy
    with cell values only, discarding all drawings, charts, and images.
    Preserves all sheets.
    """
    import tempfile

    import openpyxl

    src_wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)

    dst_wb = openpyxl.Workbook()
    # Remove default sheet created by Workbook()
    dst_wb.remove(dst_wb.active)

    for sheet_name in src_wb.sheetnames:
        src_ws = src_wb[sheet_name]
        dst_ws = dst_wb.create_sheet(title=sheet_name)

        for row in src_ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)

    clean = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    dst_wb.save(clean.name)
    clean.close()
    src_wb.close()

    return clean.name


def _save_uploaded_file(request):
    """Save uploaded file to temp, preprocess, return (clean_path, filename) or error string."""
    import os
    import tempfile

    uploaded = request.FILES["file"]
    filename = uploaded.name

    if uploaded.size > MAX_IMPORT_FILE_SIZE:
        return None, filename, f"파일 크기가 5MB를 초과합니다 ({uploaded.size / (1024*1024):.1f}MB)."

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=filename)
    for chunk in uploaded.chunks():
        tmp.write(chunk)
    tmp.close()

    try:
        clean_path = _preprocess_excel(tmp.name)
    except Exception:
        clean_path = tmp.name
    finally:
        if clean_path != tmp.name:
            try:
                os.unlink(tmp.name)
            except OSError:
                pass

    return clean_path, filename, None


def _read_raw_rows(ws, max_rows=MAX_IMPORT_ROWS):
    """Read raw rows from worksheet. Returns list of lists (values only)."""
    all_rows = []
    for row in ws.iter_rows(max_row=max_rows + 10, values_only=True):
        all_rows.append([str(c).strip() if c is not None else "" for c in row])
    return all_rows


def _count_data_rows(raw_rows, start_idx=0):
    """Count non-empty rows from start_idx."""
    count = 0
    for row in raw_rows[start_idx:]:
        if any(c for c in row):
            count += 1
    return count


def _get_max_cols(raw_rows):
    """Get max number of non-empty columns across rows."""
    return max((sum(1 for c in row if c) for row in raw_rows), default=0)


def _parse_sheet_basic(ws, max_rows=MAX_IMPORT_ROWS):
    """Quick parse for sheet classification (no AI). Returns (first_rows, total_rows, max_cols)."""
    raw_rows = _read_raw_rows(ws, max_rows)
    # Find first row with data
    start = 0
    for i, row in enumerate(raw_rows):
        if sum(1 for c in row if c) >= 2:
            start = i
            break
    total = _count_data_rows(raw_rows, start)
    max_cols = _get_max_cols(raw_rows[start:start + 6])
    first_rows = raw_rows[start:start + 6]
    return first_rows, total, max_cols


def _import_analyze(request):
    """Step 2: Parse file. Multi-sheet → sheet selection, single sheet → column mapping."""
    import openpyxl

    clean_path, filename, error = _save_uploaded_file(request)
    if error:
        return render(request, "contacts/import.html", {"error": error})

    wb = openpyxl.load_workbook(clean_path, read_only=True)
    sheet_names = wb.sheetnames

    # Single sheet → go straight to column mapping
    if len(sheet_names) == 1:
        wb.close()
        return _analyze_single_sheet(request, clean_path, filename, sheet_names=[sheet_names[0]])

    # Multiple sheets → AI classification
    sheets_info = []
    for name in sheet_names:
        ws = wb[name]
        first_rows, total_rows, max_cols = _parse_sheet_basic(ws)
        if total_rows == 0:
            sheets_info.append({
                "name": name, "headers": [], "sample": [], "rows": 0,
            })
        else:
            # Use first row as preview headers (may or may not be actual headers)
            preview_headers = [c for c in first_rows[0] if c][:10] if first_rows else []
            sheets_info.append({
                "name": name,
                "headers": preview_headers,
                "sample": [dict(zip(preview_headers, row)) for row in first_rows[1:3]] if len(first_rows) > 1 else [],
                "rows": total_rows,
            })
    wb.close()

    # AI classify sheets
    from intelligence.services import classify_sheets

    try:
        classifications = classify_sheets(sheets_info)
    except Exception:
        # Fallback: treat all non-empty sheets as contacts
        classifications = [
            {"name": s["name"], "is_contact": s["rows"] > 0, "reason": "자동 판단 실패 — 데이터 있음"}
            for s in sheets_info
        ]

    # Build display data
    sheet_results = []
    contact_sheets = []
    total_contact_rows = 0
    for info in sheets_info:
        cls = next((c for c in classifications if c["name"] == info["name"]), None)
        is_contact = cls["is_contact"] if cls else False
        reason = cls["reason"] if cls else ""
        sheet_results.append({
            "name": info["name"],
            "rows": info["rows"],
            "headers": info["headers"][:5],
            "is_contact": is_contact,
            "reason": reason,
        })
        if is_contact:
            contact_sheets.append(info["name"])
            total_contact_rows += info["rows"]

    # Store in session
    request.session["import_file"] = clean_path
    request.session["import_filename"] = filename
    request.session["import_sheet_results"] = sheet_results

    return render(request, "contacts/import_sheets.html", {
        "filename": filename,
        "sheet_results": sheet_results,
        "contact_sheets": contact_sheets,
        "total_contact_rows": total_contact_rows,
        "total_sheets": len(sheet_names),
    })


def _analyze_single_sheet(request, clean_path, filename, sheet_names):
    """Analyze selected sheet(s) with AI header detection + column mapping."""
    import os
    import openpyxl

    wb = openpyxl.load_workbook(clean_path, read_only=True)

    # Read raw data from first selected sheet (for AI analysis)
    first_sheet = None
    all_first_rows = []
    merged_total_rows = 0

    for name in sheet_names:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        first_rows, total_rows, max_cols = _parse_sheet_basic(ws)
        if total_rows == 0:
            continue
        if first_sheet is None:
            first_sheet = name
            all_first_rows = first_rows
        merged_total_rows += total_rows

    wb.close()

    if not all_first_rows:
        return render(request, "contacts/import.html", {
            "error": "선택한 시트에 데이터가 없습니다.",
        })

    if merged_total_rows > MAX_IMPORT_ROWS:
        try:
            os.unlink(clean_path)
        except OSError:
            pass
        return render(request, "contacts/import.html", {
            "error": f"데이터가 {merged_total_rows:,}건입니다. 최대 {MAX_IMPORT_ROWS:,}건까지 임포트 가능합니다.",
        })

    # AI: detect header + map columns in one call
    from intelligence.services import detect_header_and_map

    try:
        result = detect_header_and_map(all_first_rows, len(all_first_rows[0]) if all_first_rows else 0)
    except Exception:
        result = {"has_header": False, "mapping": {}}

    has_header = result["has_header"]
    mapping = result["mapping"]

    if has_header:
        # First row is header
        headers = [c for c in all_first_rows[0] if c]
        header_idx = 0
        data_start = 1
        # Re-key mapping to use header text (AI may have returned either format)
        final_mapping = {}
        for key, field in mapping.items():
            if key.startswith("col_"):
                idx = int(key.split("_")[1])
                if idx < len(headers):
                    final_mapping[headers[idx]] = field
            else:
                final_mapping[key] = field
        mapping = final_mapping
    else:
        # No header — generate col_0, col_1, ... as keys
        max_cols = max(len(r) for r in all_first_rows) if all_first_rows else 0
        headers = [f"col_{i}" for i in range(max_cols)]
        header_idx = -1  # sentinel: no header row
        data_start = 0

    # Build sample rows for preview
    sample_rows = []
    for row in all_first_rows[data_start : data_start + 4]:
        row_dict = {}
        for j, h in enumerate(headers):
            val = row[j] if j < len(row) else ""
            row_dict[h] = str(val)[:50] if val else ""
        sample_rows.append(row_dict)

    # Build preview with mapped data
    preview_rows = []
    for sample in sample_rows:
        mapped = {}
        for col, field in mapping.items():
            if field and field != "skip" and col in sample:
                if field in mapped and mapped[field]:
                    mapped[field] += f" | {sample[col]}"
                else:
                    mapped[field] = sample[col]
        preview_rows.append(mapped)

    # Store in session
    request.session["import_mapping"] = mapping
    request.session["import_file"] = clean_path
    request.session["import_header_idx"] = header_idx
    request.session["import_total"] = merged_total_rows
    request.session["import_headers"] = headers
    request.session["import_sheets"] = sheet_names
    request.session["import_has_header"] = has_header

    return render(request, "contacts/import_preview.html", {
        "mapping": mapping,
        "headers": headers,
        "preview_rows": preview_rows,
        "sample_rows": sample_rows,
        "total_rows": merged_total_rows,
        "filename": filename,
        "sheet_names": sheet_names,
        "has_header": has_header,
    })


@login_required
def contact_import_sheets(request):
    """Step 2b: User selects sheets after AI classification."""
    if request.method != "POST":
        return redirect("contacts:import")

    clean_path = request.session.get("import_file")
    filename = request.session.get("import_filename", "")

    if not clean_path:
        return redirect("contacts:import")

    action = request.POST.get("action")  # "merge" or "select"

    if action == "merge":
        # Use AI-recommended contact sheets
        sheet_results = request.session.get("import_sheet_results", [])
        selected = [s["name"] for s in sheet_results if s["is_contact"]]
    else:
        # User manually selected sheets
        selected = request.POST.getlist("sheets")

    if not selected:
        return render(request, "contacts/import.html", {
            "error": "시트를 선택해주세요.",
        })

    # Clean up session keys from sheet selection step
    request.session.pop("import_sheet_results", None)
    request.session.pop("import_filename", None)

    return _analyze_single_sheet(request, clean_path, filename, sheet_names=selected)


@login_required
def contact_import_confirm(request):
    """Step 3: Apply mapping and create contacts."""
    import openpyxl

    if request.method != "POST":
        return redirect("contacts:import")

    mapping = request.session.get("import_mapping", {})
    file_path = request.session.get("import_file")
    headers = request.session.get("import_headers", [])

    if not file_path or not mapping:
        return redirect("contacts:import")

    # Allow user overrides from form
    for h in headers:
        override = request.POST.get(f"map_{h}")
        if override is not None:
            mapping[h] = override

    has_header = request.session.get("import_has_header", True)
    wb = openpyxl.load_workbook(file_path, read_only=True)
    import_sheets = request.session.get("import_sheets")

    # Determine data start row: skip header if present
    data_skip = 1 if has_header else 0

    # Collect data rows from selected sheets (or active sheet)
    data_rows = []
    if import_sheets and len(import_sheets) >= 1:
        for sheet_name in import_sheets:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            raw_rows = _read_raw_rows(ws)
            # Find first row with data
            start = 0
            for i, row in enumerate(raw_rows):
                if sum(1 for c in row if c) >= 2:
                    start = i
                    break
            for row in raw_rows[start + data_skip :]:
                if any(c for c in row):
                    data_rows.append(row)
    else:
        ws = wb.active
        raw_rows = _read_raw_rows(ws)
        start = 0
        for i, row in enumerate(raw_rows):
            if sum(1 for c in row if c) >= 2:
                start = i
                break
        for row in raw_rows[start + data_skip :]:
            if any(c for c in row):
                data_rows.append(row)

    created = 0
    skipped = 0
    errors = 0
    pending_sentiments = []  # [(interaction_id, memo_text), ...]

    for row in data_rows:

        try:
            mapped = {}
            for j, h in enumerate(headers):
                field = mapping.get(h)
                if not field or field == "skip":
                    continue
                val = str(row[j]).strip() if j < len(row) and row[j] else ""
                if not val:
                    continue
                if field == "memo":
                    # Include column label for traceability (skip col_N labels)
                    labeled = val if h.startswith("col_") else f"{h}: {val}"
                    if field in mapped:
                        mapped[field] += f" | {labeled}"
                    else:
                        mapped[field] = labeled
                else:
                    mapped[field] = val

            name = mapped.get("name", "").strip()
            if not name:
                errors += 1
                continue

            phone = mapped.get("phone", "").strip()
            if phone and Contact.objects.filter(fc=request.user, phone=phone).exists():
                skipped += 1
                continue

            contact = Contact.objects.create(
                fc=request.user,
                name=name,
                phone=phone,
                company_name=mapped.get("company_name", ""),
                industry=mapped.get("industry", ""),
                region=mapped.get("region", ""),
                revenue_range=mapped.get("revenue_range", ""),
                employee_count=int(mapped["employee_count"]) if mapped.get("employee_count", "").isdigit() else None,
                memo=mapped.get("memo", ""),
            )

            # Create meeting if date is present
            meeting_date = mapped.get("meeting_date", "").strip()
            if meeting_date:
                from datetime import datetime, timedelta
                from meetings.models import Meeting

                try:
                    # Parse various date formats
                    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"]:
                        try:
                            dt = datetime.strptime(meeting_date.split(" ")[0][:10], fmt.split(" ")[0])
                            break
                        except ValueError:
                            continue
                    else:
                        dt = None

                    if dt:
                        from django.utils import timezone as tz
                        scheduled = tz.make_aware(dt.replace(hour=10))
                        Meeting.objects.get_or_create(
                            fc=request.user,
                            contact=contact,
                            scheduled_at=scheduled,
                            defaults={
                                "title": f"{name}님 미팅",
                                "scheduled_end_at": scheduled + timedelta(hours=1),
                                "location": mapped.get("region", ""),
                                "status": Meeting.Status.COMPLETED,
                            },
                        )
                        # Update last_interaction_at to meeting date
                        if not contact.last_interaction_at or scheduled > contact.last_interaction_at:
                            contact.last_interaction_at = scheduled
                            contact.save(update_fields=["last_interaction_at"])
                except Exception:
                    pass

            # Create interaction from memo
            memo_text = mapped.get("memo", "").strip()
            if memo_text:
                interaction = Interaction.objects.create(
                    fc=request.user,
                    contact=contact,
                    type=Interaction.Type.MEMO,
                    summary=memo_text,
                )
                pending_sentiments.append({
                    "id": str(interaction.pk),
                    "text": memo_text,
                })
                if not contact.last_interaction_at:
                    contact.last_interaction_at = interaction.created_at
                    contact.save(update_fields=["last_interaction_at"])

            created += 1
        except Exception:
            errors += 1

    # === Synchronous: Calculate relationship scores (Python, no API, <1ms/contact) ===
    from intelligence.services import calculate_relationship_score

    all_fc_contacts = list(Contact.objects.filter(fc=request.user))
    for c in all_fc_contacts:
        calculate_relationship_score(c)

    # === Collect IDs for background processing ===
    new_contact_ids = [str(c.pk) for c in Contact.objects.filter(fc=request.user).order_by("-created_at")[:created]]
    new_interaction_ids = [entry["id"] for entry in pending_sentiments]

    # === Create ImportBatch ===
    from intelligence.models import ImportBatch

    batch = ImportBatch.objects.create(
        fc=request.user,
        contact_count=created,
        interaction_count=len(new_interaction_ids),
    )

    # Link interactions to this batch
    if new_interaction_ids:
        Interaction.objects.filter(pk__in=new_interaction_ids).update(import_batch=batch)

    # === Background analysis pipeline (Gemini embedding + sentiment + task detection) ===
    if created > 0 or new_interaction_ids:
        import threading

        def _run_import_analysis(batch_id, contact_ids, interaction_ids):
            import logging
            import traceback

            from django.db import connection

            logger = logging.getLogger(__name__)
            try:
                from common.embedding import get_embeddings_batch
                from intelligence.models import ImportBatch as BgBatch
                from intelligence.services import (
                    classify_sentiments_batch,
                    detect_tasks_batch,
                    embed_contacts_batch,
                )

                bg_batch = BgBatch.objects.get(id=batch_id)
                bg_contacts = list(Contact.objects.filter(id__in=contact_ids))
                bg_interactions = list(Interaction.objects.filter(id__in=interaction_ids))

                # Step 1: Batch embed contacts (Gemini API, 100-chunk)
                embed_contacts_batch(bg_contacts)
                bg_batch.embedding_done = True
                bg_batch.save(update_fields=["embedding_done"])

                # Step 2: Generate interaction embeddings once, share for sentiment + task
                if bg_interactions:
                    texts = [i.summary for i in bg_interactions]
                    interaction_embeddings = get_embeddings_batch(texts)

                    # Step 3: Sentiment classification (embedding reuse)
                    classify_sentiments_batch(bg_interactions, embeddings=interaction_embeddings)
                    bg_batch.sentiment_done = True
                    bg_batch.save(update_fields=["sentiment_done"])

                    # Step 4: Task detection (embedding reuse)
                    detect_tasks_batch(bg_interactions, embeddings=interaction_embeddings)
                    bg_batch.task_done = True
                    bg_batch.save(update_fields=["task_done"])
                else:
                    bg_batch.sentiment_done = True
                    bg_batch.task_done = True
                    bg_batch.save(update_fields=["sentiment_done", "task_done"])

                # Step 5: Recalculate scores with sentiment data
                for c in bg_contacts:
                    calculate_relationship_score(c)
            except Exception:
                logger.exception("Import analysis pipeline failed for batch %s", batch_id)
                BgBatch.objects.filter(id=batch_id).update(
                    error_message=f"Pipeline failed: {traceback.format_exc()[:500]}"
                )
            finally:
                connection.close()

        thread = threading.Thread(
            target=_run_import_analysis,
            args=(batch.id, new_contact_ids, new_interaction_ids),
            daemon=False,
        )
        thread.start()

    # Cleanup
    import os

    try:
        os.unlink(file_path)
    except OSError:
        pass
    for key in ["import_mapping", "import_file", "import_header_idx", "import_total", "import_headers", "import_sheets", "import_has_header"]:
        request.session.pop(key, None)

    # Tier distribution for import result
    tier_counts = {}
    for c in all_fc_contacts:
        tier = c.relationship_tier or "gray"
        tier_counts[tier] = tier_counts.get(tier, 0) + 1

    return render(request, "contacts/import_result.html", {
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "total": created + skipped + errors,
        "batch": batch,
        "tier_counts": tier_counts,
    })


@login_required
def interaction_create(request, contact_pk):
    contact = get_object_or_404(Contact, pk=contact_pk, fc=request.user)

    if request.method == "POST":
        interaction = Interaction.objects.create(
            fc=request.user,
            contact=contact,
            meeting_id=request.POST.get("meeting_id") or None,
            type=request.POST.get("type", Interaction.Type.MEMO),
            summary=request.POST["summary"],
            sentiment=request.POST.get("sentiment", ""),
        )
        contact.last_interaction_at = timezone.now()
        contact.save(update_fields=["last_interaction_at"])

        # === Auto-analysis pipeline (embedding → sentiment → task → score) ===
        from intelligence.services import (
            calculate_relationship_score,
            classify_sentiment,
            detect_task,
            embed_contact,
        )

        from common.embedding import get_embedding

        task_created = False

        # Embedding once → reuse for sentiment + task detection
        embedding = get_embedding(interaction.summary)

        if embedding is not None:
            # Sentiment classification (single)
            if not interaction.sentiment:
                sentiment = classify_sentiment(interaction.summary, embedding=embedding)
                if sentiment:
                    interaction.sentiment = sentiment
                    interaction.save(update_fields=["sentiment"])

            # Task detection (single, sets task_checked=True)
            detected_task = detect_task(interaction, embedding=embedding)
            task_created = detected_task is not None

        # Score recalculation (Python, always succeeds)
        calculate_relationship_score(contact)

        # Re-embed contact (source_hash comparison, updates only if changed)
        embed_contact(contact)

        if request.htmx:
            interactions = contact.interactions.all()[:20]
            return render(request, "contacts/partials/interaction_timeline.html", {
                "contact": contact,
                "interactions": interactions,
                "task_created": task_created,
            })
        return redirect("contacts:detail", pk=contact.pk)

    return render(request, "contacts/partials/interaction_form.html", {
        "contact": contact,
        "meeting_id": request.GET.get("meeting"),
    })


@login_required
def interaction_edit(request, contact_pk, pk):
    contact = get_object_or_404(Contact, pk=contact_pk, fc=request.user)
    interaction = get_object_or_404(Interaction, pk=pk, contact=contact, fc=request.user)

    if request.method == "POST":
        interaction.summary = request.POST["summary"]
        interaction.type = request.POST.get("type", interaction.type)
        interaction.sentiment = request.POST.get("sentiment", interaction.sentiment)
        interaction.save(update_fields=["summary", "type", "sentiment"])

        if request.htmx:
            interactions = contact.interactions.all()[:INTERACTION_PAGE_SIZE]
            return render(request, "contacts/partials/interaction_timeline.html", {
                "contact": contact,
                "interactions": interactions,
            })
        return redirect("contacts:detail", pk=contact.pk)

    return render(request, "contacts/partials/interaction_edit_form.html", {
        "contact": contact,
        "interaction": interaction,
    })


@login_required
def interaction_delete(request, contact_pk, pk):
    contact = get_object_or_404(Contact, pk=contact_pk, fc=request.user)
    interaction = get_object_or_404(Interaction, pk=pk, contact=contact, fc=request.user)

    if request.method == "POST":
        interaction.delete()

        from intelligence.services import calculate_relationship_score
        calculate_relationship_score(contact)

        if request.htmx:
            interactions = contact.interactions.all()[:INTERACTION_PAGE_SIZE]
            return render(request, "contacts/partials/interaction_timeline.html", {
                "contact": contact,
                "interactions": interactions,
            })
        return redirect("contacts:detail", pk=contact.pk)

    return HttpResponse(status=405)


@login_required
def contact_ai_section(request, pk):
    """HTMX lazy-load endpoint: AI analysis for contact detail page.

    Runs ensure_embedding → ensure_sentiments_and_tasks → find_similar sequentially.
    All failures are safe — returns partial with whatever data is available.
    """
    contact = get_object_or_404(Contact, pk=pk, fc=request.user)

    from intelligence.services import (
        ensure_embedding,
        ensure_sentiments_and_tasks,
        find_similar_contacts,
    )

    # Step 1: Ensure embedding exists (Gemini API, timeout 3s)
    ensure_embedding(contact)

    # Step 2: Sentiment + task analysis for unprocessed interactions
    ensure_sentiments_and_tasks(contact)

    # Step 3: Find similar contacts (pgvector, <10ms)
    similar = find_similar_contacts(contact, n=3)

    return render(request, "contacts/partials/contact_ai_section.html", {
        "contact": contact,
        "similar_contacts": similar,
    })


@login_required
def task_create(request):
    if request.method == "POST":
        contact_id = request.POST.get("contact_id") or None
        Task.objects.create(
            fc=request.user,
            contact_id=contact_id,
            title=request.POST["title"],
            due_date=request.POST.get("due_date") or None,
        )
        # Return updated task list
        pending_tasks = Task.objects.filter(fc=request.user, is_completed=False).select_related("contact")[:10]
        return render(request, "contacts/partials/task_list_items.html", {"pending_tasks": pending_tasks})

    # GET: return inline form
    return render(request, "contacts/partials/task_form.html")


@login_required
def task_complete(request, pk):
    task = get_object_or_404(Task, pk=pk, fc=request.user)
    task.is_completed = True
    task.save(update_fields=["is_completed"])
    return HttpResponse("")  # Remove the element via hx-swap="outerHTML"


@login_required
def task_edit(request, pk):
    task = get_object_or_404(Task, pk=pk, fc=request.user)

    if request.method == "POST":
        task.title = request.POST["title"]
        task.due_date = request.POST.get("due_date") or None
        task.save(update_fields=["title", "due_date"])

        pending_tasks = Task.objects.filter(fc=request.user, is_completed=False).select_related("contact")[:10]
        return render(request, "contacts/partials/task_list_items.html", {"pending_tasks": pending_tasks})

    return render(request, "contacts/partials/task_edit_form.html", {"task": task})


@login_required
def task_delete(request, pk):
    task = get_object_or_404(Task, pk=pk, fc=request.user)

    if request.method == "POST":
        task.delete()
        if request.htmx:
            return HttpResponse("")  # Remove the element
        return redirect("home")

    return HttpResponse(status=405)
